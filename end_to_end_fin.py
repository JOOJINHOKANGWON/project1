import cv2
import numpy as np
import argparse
import os
from plantcv import plantcv as pcv
import tkinter
from PIL import Image
from tkinter import filedialog
window = tkinter.Tk()
import pandas as pd
from pandas import DataFrame

import openpyxl
###########################
#       click region      #
###########################

# just click object region like rectangle

Point_x=[]
Point_y=[]                  
isDragging = False            
count=0
# 화면 출력
colors = [] 
selected_roi=[]
Cursor_xmin,Cursor_xmax,Cursor_ymin,Cursor_ymax=0,0,0,0
rows,cols=0,0


def onMouse(event, Cursor_x, Cursor_y, flags,img):
    global Cursor_xmin,Cursor_xmax,Cursor_ymin,Cursor_ymax
    global isDragging, count ,ptsx , ptsy, selected_roi
    
    if event==cv2.EVENT_LBUTTONDOWN:
        isDragging=True
        
    elif event==cv2.EVENT_MOUSEMOVE:
        if isDragging:
        
            cv2.imshow('watershed',img)

    elif event==cv2.EVENT_LBUTTONUP:
        isDragging=False
        Point_x.append(Cursor_x)
        Point_y.append(Cursor_y)
        print(Point_x)     #
        print(Point_y)     #
        Cursor_xmin=min(Point_x)     #
        Cursor_xmax=max(Point_x)     #                   점으로 찍은 네 좌표를 취합해서
        Cursor_ymin=min(Point_y)     #                  직사각형을 그림
        Cursor_ymax=max(Point_y)
        count+=1  #
        print("count :",count)    #
        if count==4:    #
            selected_roi=img[Cursor_ymin:Cursor_ymax,Cursor_xmin:Cursor_xmax]
            
            print(Cursor_ymin,Cursor_ymax,Cursor_xmin,Cursor_xmax)
            
            cv2.destroyAllWindows()


def Start_A(img):
    global count
    global rows,cols
    global Cursor_xmin,Cursor_xmax,Cursor_ymin,Cursor_ymax
    rows, cols = img.shape[:2]
    
    cv2.imshow('watershed',img)
    while True:
        cv2.setMouseCallback('watershed', onMouse,img)
        cv2.waitKey(0)
        if count==4:
            break
    return selected_roi,Cursor_xmin,Cursor_xmax,Cursor_ymin,Cursor_ymax

def skeleton(img): ## findnode 함수에 넣기 위해 이미지를 스켈레톤화 함
    img=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
    size = np.size(img)
    skel = np.zeros(img.shape,np.uint8)

    _,img = cv2.threshold(img,127,255,0)
    element = cv2.getStructuringElement(cv2.MORPH_ELLIPSE,(3,3))
    done = False

    img =cv2.bitwise_not(img)
    
    while( not done):
        eroded = cv2.erode(img,element)  ## 스켈레톤화 과정  침식 팽창연산을 순서대로 진행하여 잡음을 전부 지우고
        temp = cv2.dilate(eroded,element)  ## 원본 이미지에서 잡음을 지운 이미지를 뺌
        temp = cv2.subtract(img,temp)   ## 그렇게 하면 남아있는 것들은 모두 0이 되고 잡음들만 원래의 값을 가지고 있음 (temp 변수에)
        skel = cv2.bitwise_or(skel,temp) ## temp 변수와 0으로 채워진 이미지와 비트연산으로 skel에 계속 누적함 >> skel 이미지에 잡음들이 모여 skeletonize
        img = eroded.copy()

        zeros = size - cv2.countNonZero(img)
        if zeros==size:
            done = True

    return skel

## 스켈레톤화 한 이미지로 노드를 찍은 이미지 출력## 
## 현재 노드를 찍은 이미지에서 노드의 좌표를 읽어 올 방법 고안중##  ## 2021 05. 해결
def findnode(img):
    ret,img=cv2.threshold(img,127,255,0)
    skeleton=pcv.morphology.skeletonize(mask=img)

    branch_points_img=pcv.morphology.find_branch_pts(skel_img=skeleton,mask=None) ## 현재 pcv 함수에서 저절로 이미지가 뜨는데 이걸 찾아서 지워야함 (06.10)
                                                                                
    return branch_points_img

def find(img6):  ### findnode에서 넘긴 이미지에서 함수가 노드라고 인식한 점들의 좌표를 리스트에 저장
    print(img6)
    sum=0
    node_x_point=[]
    node_y_point=[]
    rows,cols=img6.shape[:2]
    for row in range(rows):
        for col in range(cols):
            if img6[row,col] ==0:
                continue
            node_y_point.append(row)
            node_x_point.append(col)
            sum+=1
    return node_x_point,node_y_point

def sobel(img,imgreal,Cursor_xmin,Cursor_xmax,Cursor_ymin,Cursor_ymax,node_x_point,node_y_point,tip_x_point,tip_y_point):
    selected_roi_copy=img.copy()
    
    
    tips_x_max=max(tip_x_point)
    tips_y_min=min(tip_y_point)
    tips_y_max=max(tip_y_point)
    tips_x_min=min(tip_x_point)
        
        
    node_x_max=max(node_x_point)  ## 가장 낮은 노드의 x 좌표
    
    node_y_max=max(node_y_point)  ## 가장 낮은 노드의 y 좌표
    
    
    
    
    ##cv2.rectangle(imgreal,(Cursor_xmin+tips_x_min,Cursor_ymin+tips_y_min),(Cursor_xmin+node_x_max,Cursor_ymin+tips_y_max),(0,0,255),thickness=3)   ##전체높이 영역 바운딩
    ##cv2.rectangle(imgreal,(Cursor_xmin+tips_x_min,Cursor_ymin+tips_y_min),(Cursor_xmin+node_x_max,Cursor_ymin+node_y_max),(255,0,0),thickness=2)  ##수관높이 영역 바운딩
    ##cv2.rectangle(imgreal,(Cursor_xmin+tips_x_min,Cursor_ymin+node_y_max),(Cursor_xmin+node_x_max,Cursor_ymin+tips_y_max),(0,255,255)) ##수간높이 영역 바운딩
    return selected_roi_copy,imgreal,node_x_max-tips_x_min,tips_y_max-tips_y_min,tips_y_max-node_y_max


def tracking(bounding_image):
    
    lower_blue=np.array([150,100,90])
    upper_blue=np.array([225,225,255])

    frame23=bounding_image
        
    hsv=cv2.cvtColor(frame23,cv2.COLOR_BGR2HSV)
    blue_range=cv2.inRange(hsv,lower_blue,upper_blue)    ## 마스크 영역 분할 >> 위의 lower_blue, upper_blue 사이의 값을 제외하고는 전부 0으로 찍힘
    
    blue_result=blue_range
    try:
        _,contours1,_=cv2.findContours(blue_result,cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)
    except:
        contours1,_=cv2.findContours(blue_result,cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)
    for contour in contours1:
        (x,y,w,h)=cv2.boundingRect(contour)
        if w<20:      ## 작게 잡히는 노이즈를 제거
            continue
        if h<20:      ##  위와 같음
            continue
        pixel_width=w     ## 실제 길이를 구하기위해 checkparm에서 사용 할 마스크 너비
        pixel_height=h     ## 실제 길이를 구하기위해 checkparm에서 사용 할 마스크 높이
        
    ##cv2.rectangle(bounding_image,pt1=(x,y),pt2=(x+w,y+h),color=(255,0,255),thickness=3) ##분홍색 마스크 영역을 바운딩함

    return bounding_image,pixel_height,pixel_width


mask_real_height=0  ## 실제 마스크 높이
mask_real_width=0  ## 실제 마스크 너비
def checkhow(h,w,object_width,object_height,object_spec_height):

    print('height maskpixel :',h)  ## 마스크 높이 픽셀값
    print("width mask pixel:", w)  ## 마스크 너비 픽셀값
    print('object width :', object_width) ## 나무 폭
    print('object height ', object_height) ## 나무 전체높이
    print('object spec height', object_spec_height) ## 나무 수간높이

    real_height=mask_real_height/h ## 실제 길이를 마스크 픽셀길이로 나눠서 1픽셀당 실제 길이를 realh라는 변수로 잡음
    real_width=mask_real_width/w ## 위와 똑같음 realw는 1픽셀당 실제 너비
    

    final_height=int(real_height*object_height)  ##전체높이
    final_width=int(real_width*object_width)  ## 전체 폭
    smh=int(real_height*object_spec_height)     ## 수간높이
    sgh=int(final_height-smh)   ## 수관높이
    print("수목높이,수목너비,수간높이,수관높이",final_height,final_width,smh,sgh)
    return final_height,final_width,smh,sgh

def findtips(img):
    
    
    ret,img=cv2.threshold(img,127,255,0)
    skeleton=pcv.morphology.skeletonize(mask=img)
    

    pcv.params.line_thickness=3


    tips_img=pcv.morphology.find_tips(skel_img=skeleton)
    
    return tips_img

def calc(event):
    global label,entry,mask_real_height
    mask_real_height=eval(entry.get())
    label.config(text="마스크 높이:" +str(eval(entry.get())))


def calc2(event):
    global label1,entry,mask_real_width
    mask_real_width=eval(entry2.get())
    label1.config(text="마스크 너비:" +str(eval(entry2.get())))
    cv2.cvtColor(image,cv2.BGR2)

label3=tkinter.Label(window,text="엔터를 누르면 값이 저장됩니다",width=50,height=10)
entry=tkinter.Entry(window,width=30)
label=tkinter.Label(window,text="마스크 높이",width=50,height=10)

entry2=tkinter.Entry(window,width=30)
label1=tkinter.Label(window,text="마스크 너비",width=30,height=10)



def main2():
    global entry, entry2
    window.title("화난다")
    
    window.geometry("1080x1080+100+100")

    btn_open=tkinter.Button(window,text="파일 경로 설정",command=find_dir)

    btn= tkinter.Button(window,text='시작',command=main)
    
    btn_open.pack()

    entry.bind("<Return>",calc)
    label3.pack()
    label.pack()
    entry.pack()

    entry2.bind("<Return>",calc2)
    label1.pack()
    entry2.pack()

    btn.pack()

    window.resizable(False,False)
    window.mainloop()


file_path=""
def find_dir():
    global file_path
    file_path=tkinter.filedialog.askdirectory(parent=window,initialdir="/",title='please select a directory')
    


    


def main(): 
  global count,file_path
  global Point_x, Point_y
  
  wb=openpyxl.Workbook()
  ws=wb.active
  num=2
  ws.cell(row=1,column=2).value="수목높이"
  ws.cell(row=1,column=3).value="수목너비"
  ws.cell(row=1,column=4).value="수간높이"
  ws.cell(row=1,column=5).value="수관높이"
  lst_dir=[]
  
  
  
  lst_dir=os.listdir(f'{file_path}')
  for i in lst_dir:
    Point_x=[]
    Point_y=[]
    Img_origin=cv2.imread(f'{file_path}\\'+i)
  
    count=0;
    rows,cols=Img_origin.shape[:2]
    if rows>cols:
            Img_origin=cv2.resize(Img_origin,(int(1024*(rows/cols)),1024))
    elif rows<cols:
            Img_origin=cv2.resize(Img_origin,(1024,int(1024*(rows/cols))))
    else:
            Img_origin=cv2.resize(Img_origin,(1024,1024))
    rows,cols=Img_origin.shape[:2]
    ##ones=cv2.imread('C:\\Users\\pllab\\Desktop\\wjdrma\\dataset\\white.png')
    ##ones=cv2.resize(ones,(cols,rows))
    ##Img_origin=cv2.addWeighted(Img_origin,float(80)*0.01,ones,float(20)*0.01,0)
    img_origin_copy=Img_origin.copy()
    selected_roi,Cursor_xmin,Cursor_xmax,Cursor_ymin,Cursor_ymax=Start_A(Img_origin) ## 보낸 이미지에서 물체를 분할하기 위해 사용자가 임의로 영역을 잡음. 이는 후에 딥러닝으로 바꿀 예정
    roi_skeletonize=skeleton(selected_roi) ## findnode에서 이미지를 돌리기 위해서 스켈레톤화를 진행함
    tip_binary_image=findtips(roi_skeletonize)
    tip_binary_numpy_image=np.array(tip_binary_image)
    tip_x_point,tip_y_point=find(tip_binary_numpy_image)
    node_binary_image=findnode(roi_skeletonize) ## plantcv의 findnode 함수를 이용해서 가장 아래 브랜치를 구함
    node_binary_numpy_image=np.array(node_binary_image) ##findxy에서 넘파이 배열을 필요로 하기 때문에 넘파이배열로 바꿔줌
    node_x_point,node_y_point=find(node_binary_numpy_image) ## plantcv에서 findnode 돌린 이미지에서 가장 바깥 영역을 찾음
    selected_roi_copy,img_origin_copy,object_width,object_height,object_spec_height=sobel(selected_roi,Img_origin,Cursor_xmin,Cursor_xmax,
    Cursor_ymin,Cursor_ymax,node_x_point,node_y_point,tip_x_point,tip_y_point) ## 영역분리 지금 엠보싱 연산으로 전체길이를 구했는데 이는 매우 부정확, plantcv의 findtip 으로 구하면 편할듯
    
    try:
        img3,h,w=tracking(img_origin_copy) ## 색깔 마스크 찾는거 분홍색으로만 해놓음 ##img3가 최종사진
        final_height,final_width,smh,sgh=checkhow(h,w,object_width,object_height,object_spec_height)  ## 실제길이 파악
        
        ws.cell(row=num,column=1).value=f'{i}'
        ws.cell(row=num,column=2).value=f'{final_height}'
        ws.cell(row=num,column=3).value=f'{final_width}'
        ws.cell(row=num,column=4).value=f'{smh}'
        ws.cell(row=num,column=5).value=f'{sgh}'
        num+=1
        
        
        
            
            
        
    except:
        print("no mask")
   
  wb.save('number_format1.xlsx')
  
  
  

main2()